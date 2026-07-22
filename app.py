r"""
Flask application for the Haier ESD Monitoring & Dashboard System.
Includes all 8 sidebar modules: Dashboard, Live Monitor, History, Reports, Search, Alarms, Settings, Users.
Adheres to standard Flask templates architecture.
"""

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import json
import queue
import random
import threading
import time
import csv
import io
import datetime
import logging
import math
from logging.handlers import RotatingFileHandler
from functools import wraps

from flask import (
    Flask,
    jsonify,
    request,
    session,
    Response,
    render_template,
    redirect,
    url_for,
    stream_with_context,
)
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler
from sqlalchemy import text, func
from sqlalchemy.orm import joinedload
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import get_config
import database as db


# ==================================================================== #
# Hardware Layer
# ==================================================================== #
class KC868Driver:
    def __init__(self, host, port, unit_id, total_channels):
        self.host = host
        self.port = port
        self.unit_id = unit_id
        self.total_channels = 20  # Explicitly restricted to 20 stations
        self._connected = False

    def connect(self):
        self._connected = True
        return True

    def read_channels(self):
        return {ch: False for ch in range(1, self.total_channels + 1)}

    def close(self):
        self._connected = False


class SimulatedKC868Driver(KC868Driver):
    def __init__(self, host, port, unit_id, total_channels):
        super().__init__(host, port, unit_id, total_channels)
        self._state = {ch: False for ch in range(1, 21)}  # Hard limited to 20 channels

    def connect(self):
        self._connected = True
        return True

    def read_channels(self):
        if not self._connected:
            raise ConnectionError("Simulated KC868 not connected")
        if random.random() < 0.05:
            ch = random.randint(1, 20)
            self._state[ch] = not self._state[ch]
        return dict(self._state)


def build_driver(cfg):
    # Enforces maximum limit of 20 channels
    return SimulatedKC868Driver(cfg.KC868_HOST, cfg.KC868_PORT, cfg.KC868_UNIT_ID, 20)


# ==================================================================== #
# Real-time push - Server-Sent Events broadcaster
# ==================================================================== #
class SSEBroadcaster:
    def __init__(self):
        self._subscribers = []
        self._lock = threading.Lock()

    def subscribe(self):
        q = queue.Queue(maxsize=100)
        with self._lock:
            self._subscribers.append(q)
        return q

    def unsubscribe(self, q):
        with self._lock:
            if q in self._subscribers:
                self._subscribers.remove(q)

    def publish(self, event_name, payload):
        data = json.dumps(payload, default=str)
        message = f"event: {event_name}\ndata: {data}\n\n"
        with self._lock:
            for q in list(self._subscribers):
                try:
                    q.put_nowait(message)
                except queue.Full:
                    pass


broadcaster = SSEBroadcaster()


# ==================================================================== #
# Service Layer - Debounce + Event creation
# ==================================================================== #
class DebounceService:
    def __init__(self, required_consecutive):
        self.required_consecutive = required_consecutive
        self._pending = {}

    def filter(self, channel, new_state, last_confirmed_state):
        if new_state == last_confirmed_state:
            self._pending.pop(channel, None)
            return False

        pending = self._pending.get(channel)
        if pending and pending["state"] == new_state:
            pending["count"] += 1
        else:
            pending = {"state": new_state, "count": 1}
        self._pending[channel] = pending

        if pending["count"] >= self.required_consecutive:
            self._pending.pop(channel, None)
            return True
        return False


class EventService:
    def __init__(self, app_config):
        self.cfg = app_config

    def record_transition(self, station, is_violation, session_ref):
        now = datetime.datetime.now(datetime.UTC).replace(tzinfo=None)
        event_type = "VIOLATION" if is_violation else "RESTORED"
        duration_seconds = None

        if not is_violation:
            last_violation = (
                session_ref.query(db.ESDEvent)
                .filter_by(station_id=station.id, event_type="VIOLATION")
                .order_by(db.ESDEvent.event_timestamp.desc())
                .first()
            )
            if last_violation:
                lv_time = last_violation.event_timestamp.replace(tzinfo=None)
                duration_seconds = int((now - lv_time).total_seconds())

        new_event = db.ESDEvent(
            station_id=station.id,
            event_type=event_type,
            event_timestamp=now,
            duration_seconds=duration_seconds,
            acknowledged=False,
        )
        session_ref.add(new_event)

        station.current_state = is_violation
        session_ref.add(station)
        session_ref.commit()

        # Count total unique alerts triggered today
        today_start = datetime.datetime.now(datetime.UTC).replace(hour=0, minute=0, second=0, microsecond=0).replace(tzinfo=None)
        today_alerts_count = session_ref.query(db.ESDEvent).filter(
            db.ESDEvent.event_type == "VIOLATION",
            db.ESDEvent.event_timestamp >= today_start
        ).count()

        broadcaster.publish(
            "esd_event",
            {
                "station_code": station.station_code,
                "station_id": station.id,
                "event_type": event_type,
                "event_timestamp": now.isoformat(),
                "duration_seconds": duration_seconds,
                "today_alerts_count": today_alerts_count,
                "violation_start": now.isoformat() if is_violation else None
            },
        )
        return new_event

    def record_comm_lost(self, session_ref):
        health = db.SystemHealth(
            check_timestamp=datetime.datetime.now(datetime.UTC).replace(tzinfo=None),
            kc868_status="OFFLINE",
            response_time_ms=None,
            notes="Watchdog: consecutive poll failures exceeded threshold",
        )
        session_ref.add(health)
        session_ref.commit()
        broadcaster.publish("comm_status", {"status": "OFFLINE", "timestamp": health.check_timestamp.isoformat()})

    def record_comm_ok(self, session_ref, response_time_ms):
        health = db.SystemHealth(
            check_timestamp=datetime.datetime.now(datetime.UTC).replace(tzinfo=None),
            kc868_status="ONLINE",
            response_time_ms=response_time_ms,
        )
        session_ref.add(health)
        session_ref.commit()
        broadcaster.publish("comm_status", {"status": "ONLINE", "timestamp": health.check_timestamp.isoformat()})


# ==================================================================== #
# Poller (background thread scheduled by APScheduler)
# ==================================================================== #
class PollerThread:
    def __init__(self, app_config, driver, event_service):
        self.cfg = app_config
        self.driver = driver
        self.event_service = event_service
        self.debounce = DebounceService(app_config.DEBOUNCE_CONSECUTIVE_POLLS)
        self._last_confirmed_state = {}
        self._consecutive_failures = 0
        self._comm_lost = False
        self.logger = logging.getLogger("hardware")

    def poll_once(self):
        session_ref = db.get_session()
        try:
            start = time.monotonic()
            channel_states = self.driver.read_channels()
            elapsed_ms = int((time.monotonic() - start) * 1000)

            if self._comm_lost:
                self._comm_lost = False
                self.event_service.record_comm_ok(session_ref, elapsed_ms)
            self._consecutive_failures = 0

            stations = session_ref.query(db.Station).filter_by(status="ACTIVE").filter(db.Station.id <= 20).all()
            stations_by_channel = {s.kc868_channel: s for s in stations}

            for channel, is_violation in channel_states.items():
                if channel > 20:
                    continue
                station = stations_by_channel.get(channel)
                if station is None:
                    continue

                last_state = self._last_confirmed_state.get(channel, False)
                if self.debounce.filter(channel, is_violation, last_state):
                    self._last_confirmed_state[channel] = is_violation
                    self.event_service.record_transition(station, is_violation, session_ref)
                    self.logger.info(
                        "Station %s channel %s transitioned to %s",
                        station.station_code, channel, "VIOLATION" if is_violation else "RESTORED",
                    )
        except Exception as exc:
            self._consecutive_failures += 1
            self.logger.warning("Poll failure (%s/%s): %s",
                                 self._consecutive_failures, self.cfg.MAX_CONSECUTIVE_POLL_FAILURES, exc)
            if self._consecutive_failures >= self.cfg.MAX_CONSECUTIVE_POLL_FAILURES and not self._comm_lost:
                self._comm_lost = True
                self.event_service.record_comm_lost(session_ref)
        finally:
            db.remove_session()


# ==================================================================== #
# Auth helpers
# ==================================================================== #
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"status": "error", "data": None, "message": "Authentication required"}), 401
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def role_required(*allowed_roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if session.get("role") not in allowed_roles:
                if request.path.startswith("/api/"):
                    return jsonify({"status": "error", "data": None, "message": "Insufficient permissions"}), 403
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


# ==================================================================== #
# Logging setup
# ==================================================================== #
def configure_logging(cfg):
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")

    app_handler = RotatingFileHandler(cfg.APP_LOG_FILE, maxBytes=cfg.LOG_MAX_BYTES, backupCount=cfg.LOG_BACKUP_COUNT)
    app_handler.setFormatter(formatter)
    app_handler.setLevel(logging.INFO)

    hw_handler = RotatingFileHandler(cfg.HARDWARE_LOG_FILE, maxBytes=cfg.LOG_MAX_BYTES, backupCount=cfg.LOG_BACKUP_COUNT)
    hw_handler.setFormatter(formatter)
    hw_handler.setLevel(logging.DEBUG)

    err_handler = RotatingFileHandler(cfg.ERROR_LOG_FILE, maxBytes=cfg.LOG_MAX_BYTES, backupCount=cfg.LOG_BACKUP_COUNT)
    err_handler.setFormatter(formatter)
    err_handler.setLevel(logging.ERROR)

    logging.getLogger("werkzeug").addHandler(app_handler)
    hw_logger = logging.getLogger("hardware")
    hw_logger.setLevel(logging.DEBUG)
    hw_logger.addHandler(hw_handler)

    root_logger = logging.getLogger()
    root_logger.addHandler(err_handler)
    root_logger.setLevel(logging.INFO)


# ==================================================================== #
# Flask Application Setup & Engine Integration
# ==================================================================== #
app = Flask(__name__)
cfg = get_config()
app.secret_key = cfg.SECRET_KEY

configure_logging(cfg)
db.init_db()


def load_db_config():
    session_ref = db.get_session()
    try:
        host_cfg = session_ref.query(db.SystemConfig).filter_by(key="KC868_HOST").first()
        port_cfg = session_ref.query(db.SystemConfig).filter_by(key="KC868_PORT").first()
        interval_cfg = session_ref.query(db.SystemConfig).filter_by(key="POLL_INTERVAL_MS").first()
        
        if not host_cfg:
            host_cfg = db.SystemConfig(key="KC868_HOST", value=cfg.KC868_HOST)
            session_ref.add(host_cfg)
        if not port_cfg:
            port_cfg = db.SystemConfig(key="KC868_PORT", value=str(cfg.KC868_PORT))
            session_ref.add(port_cfg)
        if not interval_cfg:
            interval_cfg = db.SystemConfig(key="POLL_INTERVAL_MS", value=str(cfg.POLL_INTERVAL_MS))
            session_ref.add(interval_cfg)
            
        session_ref.commit()
        
        # Apply configurations to runtime cfg
        cfg.KC868_HOST = host_cfg.value
        cfg.KC868_PORT = int(port_cfg.value)
        cfg.POLL_INTERVAL_MS = int(interval_cfg.value)
    except Exception as e:
        session_ref.rollback()
        logging.error("Failed to load configs: %s", e)
    finally:
        db.remove_session()


load_db_config()

driver = build_driver(cfg)
driver.connect()

event_service = EventService(cfg)
poller = PollerThread(cfg, driver, event_service)

scheduler = BackgroundScheduler()
scheduler.add_job(id='poll_job', func=poller.poll_once, trigger="interval", seconds=(cfg.POLL_INTERVAL_MS / 1000))
scheduler.start()


def seed_initial_records():
    session_ref = db.get_session()
    try:
        admin_user = session_ref.query(db.User).filter_by(username="admin").first()
        if not admin_user:
            admin_user = db.User(
                username="admin",
                password_hash=generate_password_hash("admin123"),
                role="ADMIN",
                full_name="Haier Admin Supervisor"
            )
            session_ref.add(admin_user)

        anas_user = session_ref.query(db.User).filter_by(username="anas").first()
        if not anas_user:
            anas_user = db.User(
                username="anas",
                password_hash=generate_password_hash("anas123"),
                role="ADMIN",
                full_name="Muhammad Anas",
                is_active=True
            )
            session_ref.add(anas_user)
        else:
            anas_user.password_hash = generate_password_hash("anas123")
            anas_user.full_name = "Muhammad Anas"
            anas_user.role = "ADMIN"
            session_ref.add(anas_user)

        # 1. Clean up events and stations > 20 to avoid integrity errors and ensure exactly 20 stations
        extra_events = session_ref.query(db.ESDEvent).filter(db.ESDEvent.station_id > 20).all()
        for ev in extra_events:
            session_ref.delete(ev)
        
        extra_stations = session_ref.query(db.Station).filter(db.Station.id > 20).all()
        for st in extra_stations:
            session_ref.delete(st)

        station_names = {
            1: "Lightrod pasting",
            2: "Masking remove",
            3: "Lightrod connections",
            4: "Reflector sheet (1)",
            5: "Reflector sheet (2)",
            6: "Diffuser Plate",
            7: "Bare frame support",
            8: "Open cell",
            9: "Inverting sides",
            10: "Screw bare frame",
            11: "Barcode online",
            12: "Wifi data cable",
            13: "Base stand holder",
            14: "Screw of base stand and holder",
            15: "Fitting of speaker cover",
            16: "Main Board",
            17: "Main board screw",
            18: "Connection (1)",
            19: "Connection (2)",
            20: "Marking"
        }

        existing_stations = {s.id: s for s in session_ref.query(db.Station).all()}
        
        for i in range(1, 21):
            desc = station_names.get(i, f"Junction Line Station {i:02d}")
            if i not in existing_stations:
                st = db.Station(
                    id=i,
                    station_code=f"ESD-STN-{i:02d}",
                    kc868_channel=i,
                    description=desc,
                    status="ACTIVE",
                    current_state=False,
                )
                session_ref.add(st)
            else:
                st = existing_stations[i]
                st.station_code = f"ESD-STN-{i:02d}"
                st.kc868_channel = i
                st.description = desc
                st.status = "ACTIVE"
                session_ref.add(st)
                    
        session_ref.commit()
    except Exception as e:
        session_ref.rollback()
        raise e
    finally:
        db.remove_session()


seed_initial_records()


# ==================================================================== #
# ROUTING CONTROLLERS
# ==================================================================== #

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        session_ref = db.get_session()
        try:
            user = session_ref.query(db.User).filter_by(username=username).first()
            if user and check_password_hash(user.password_hash, password):
                session["user_id"] = user.id
                session["username"] = user.username
                session["full_name"] = user.full_name
                session["role"] = user.role
                return redirect(url_for("dashboard"))
            return render_template("login.html", error="Invalid credentials")
        finally:
            db.remove_session()

    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", active_page="dashboard")


@app.route("/live-monitor")
@login_required
def live_monitor():
    return render_template("live_monitor.html", active_page="live_monitor")


@app.route("/history")
@login_required
def history():
    page = request.args.get("page", 1, type=int)
    limit = 15
    offset = (page - 1) * limit
    
    station_code = request.args.get("station_code")
    event_type = request.args.get("event_type")
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")
    
    session_ref = db.get_session()
    try:
        query = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(db.Station.id <= 20)
        
        if station_code:
            query = query.filter(db.Station.station_code == station_code)
        if event_type:
            query = query.filter(db.ESDEvent.event_type == event_type)
        if start_date_str:
            dt = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
            query = query.filter(db.ESDEvent.event_timestamp >= dt)
        if end_date_str:
            dt = datetime.datetime.strptime(end_date_str, "%Y-%m-%d") + datetime.timedelta(days=1)
            query = query.filter(db.ESDEvent.event_timestamp < dt)
            
        total_records = query.count()
        total_pages = math.ceil(total_records / limit)
        
        events = query.order_by(db.ESDEvent.event_timestamp.desc()).offset(offset).limit(limit).all()
        
        events_data = []
        for e in events:
            if e.event_type == "RESTORED" and e.duration_seconds is not None:
                start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                end_dt = e.event_timestamp
                start_str = start_dt.strftime('%Y-%m-%d %H:%M:%S')
                end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                start_str = e.event_timestamp.strftime('%Y-%m-%d %H:%M:%S')
                end_str = "Active" if (e.event_type == "VIOLATION" and e.station and e.station.current_state) else "--"
            
            events_data.append({
                "station_id": e.station.id if e.station else 0,
                "station_code": e.station.station_code.replace("ESD-STN-", "ST-") if e.station else "",
                "description": e.station.description if e.station else "",
                "event_type": e.event_type,
                "start_time": start_str,
                "end_time": end_str,
                "duration_seconds": e.duration_seconds,
                "acknowledged": e.acknowledged
            })
            
        return render_template(
            "history.html",
            active_page="history",
            events=events_data,
            selected_station=station_code,
            selected_event=event_type,
            start_date=start_date_str,
            end_date=end_date_str,
            current_page=page,
            total_pages=total_pages
        )
    finally:
        db.remove_session()


@app.route("/reports")
@login_required
def reports():
    range_type = request.args.get("range", "daily")
    now = datetime.datetime.now(datetime.UTC).replace(tzinfo=None)
    
    if range_type == "weekly":
        start_time = now - datetime.timedelta(days=7)
    elif range_type == "monthly":
        start_time = now - datetime.timedelta(days=30)
    else: # daily
        start_time = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
    session_ref = db.get_session()
    try:
        total_violations = session_ref.query(db.ESDEvent).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time,
            db.ESDEvent.event_type == "VIOLATION"
        ).count()
        
        avg_res = session_ref.query(func.avg(db.ESDEvent.duration_seconds)).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time,
            db.ESDEvent.event_type == "RESTORED",
            db.ESDEvent.duration_seconds != None
        ).scalar()
        avg_duration = int(avg_res) if avg_res is not None else 0
        
        top_viol = session_ref.query(
            db.ESDEvent.station_id,
            func.count(db.ESDEvent.id).label("cnt")
        ).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time,
            db.ESDEvent.event_type == "VIOLATION"
        ).group_by(db.ESDEvent.station_id).order_by(text("cnt DESC")).first()
        
        top_violator = None
        if top_viol:
            st = session_ref.get(db.Station, top_viol[0])
            if st:
                top_violator = st.station_code.replace("ESD-STN-", "ST-")
                
        # Query events within range for report table
        events = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time
        ).order_by(db.ESDEvent.event_timestamp.desc()).all()

        events_data = []
        for e in events:
            if e.event_type == "RESTORED" and e.duration_seconds is not None:
                start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                end_dt = e.event_timestamp
                start_str = start_dt.strftime('%Y-%m-%d %H:%M:%S')
                end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                start_str = e.event_timestamp.strftime('%Y-%m-%d %H:%M:%S')
                end_str = "Active" if (e.event_type == "VIOLATION" and e.station and e.station.current_state) else "--"
            
            events_data.append({
                "station_id": e.station.id if e.station else 0,
                "station_code": e.station.station_code.replace("ESD-STN-", "ST-") if e.station else "",
                "description": e.station.description if e.station else "",
                "event_type": e.event_type,
                "start_time": start_str,
                "end_time": end_str,
                "duration_seconds": e.duration_seconds,
                "acknowledged": e.acknowledged
            })
                
        return render_template(
            "reports.html",
            active_page="reports",
            selected_range=range_type,
            total_violations=total_violations,
            avg_duration_seconds=avg_duration,
            top_violator=top_violator,
            events=events_data
        )
    finally:
        db.remove_session()


@app.route("/reports/export/csv")
@login_required
def export_csv_report():
    range_type = request.args.get("range", "daily")
    now = datetime.datetime.now(datetime.UTC).replace(tzinfo=None)
    
    if range_type == "weekly":
        start_time = now - datetime.timedelta(days=7)
    elif range_type == "monthly":
        start_time = now - datetime.timedelta(days=30)
    else: # daily
        start_time = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
    session_ref = db.get_session()
    try:
        events = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time
        ).order_by(db.ESDEvent.event_timestamp.desc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESD Audit Trail"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        fill_title = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
        font_meta_lbl = Font(name="Segoe UI", size=10, bold=True, color="748297")
        font_meta_val = Font(name="Segoe UI", size=10, bold=True, color="141B2D")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_data = Font(name="Segoe UI", size=10, color="000000")
        fill_even = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "HAIER - ESD MONITORING SYSTEM"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        ws['A2'] = "Report Type:"
        ws['A2'].font = font_meta_lbl
        ws['B2'] = range_type.upper()
        ws['B2'].font = font_meta_val
        
        ws['A3'] = "Generated At:"
        ws['A3'].font = font_meta_lbl
        ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['B3'].font = font_meta_val
        
        headers = ["S.No", "Station Code", "Transition Event", "Violation Start Time", "Violation End Time", "Duration (HH:MM:SS)", "Acknowledged"]
        for col_idx, text_val in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=text_val)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[5].height = 26
        
        row_idx = 6
        for idx, e in enumerate(events, start=1):
            c1 = ws.cell(row=row_idx, column=1, value=idx)
            c1.alignment = Alignment(horizontal="center")
            
            st_code = e.station.station_code.replace("ESD-STN-", "ST-") if e.station else ""
            c2 = ws.cell(row=row_idx, column=2, value=st_code)
            c2.alignment = Alignment(horizontal="center")
            
            c3 = ws.cell(row=row_idx, column=3, value=e.event_type)
            c3.alignment = Alignment(horizontal="center")
            
            # Start and End Time logic
            if e.event_type == "RESTORED" and e.duration_seconds is not None:
                start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                end_val = e.event_timestamp.replace(tzinfo=None)
            else:
                start_dt = e.event_timestamp
                end_val = "--"
                
            c4 = ws.cell(row=row_idx, column=4, value=start_dt.replace(tzinfo=None))
            c4.number_format = 'yyyy-mm-dd hh:mm:ss'
            c4.alignment = Alignment(horizontal="center")
            
            c5 = ws.cell(row=row_idx, column=5, value=end_val)
            if isinstance(end_val, datetime.datetime):
                c5.number_format = 'yyyy-mm-dd hh:mm:ss'
            c5.alignment = Alignment(horizontal="center")
            
            dur_str = "--"
            if e.duration_seconds is not None:
                hrs = e.duration_seconds // 3600
                mins = (e.duration_seconds % 3600) // 60
                secs = e.duration_seconds % 60
                dur_str = f"{hrs:02d}:{mins:02d}:{secs:02d}"
            c6 = ws.cell(row=row_idx, column=6, value=dur_str)
            c6.alignment = Alignment(horizontal="center")
            
            ack_str = "YES" if e.acknowledged else "NO"
            c7 = ws.cell(row=row_idx, column=7, value=ack_str)
            c7.alignment = Alignment(horizontal="center")
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_even
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime.datetime):
                        val_len = 19
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        ws.freeze_panes = 'A6'
        
        if row_idx > 6:
            ws.auto_filter.ref = f"A5:G{row_idx - 1}"
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename=esd_report_{range_type}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return response
    finally:
        db.remove_session()


@app.route("/reports/export/excel")
@login_required
def export_excel_report():
    range_type = request.args.get("range", "daily")
    now = datetime.datetime.now(datetime.UTC).replace(tzinfo=None)
    
    if range_type == "weekly":
        start_time = now - datetime.timedelta(days=7)
    elif range_type == "monthly":
        start_time = now - datetime.timedelta(days=30)
    else: # daily
        start_time = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
    session_ref = db.get_session()
    try:
        events = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_timestamp >= start_time
        ).order_by(db.ESDEvent.event_timestamp.desc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESD Audit Trail"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        fill_title = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
        font_meta_lbl = Font(name="Segoe UI", size=10, bold=True, color="748297")
        font_meta_val = Font(name="Segoe UI", size=10, bold=True, color="141B2D")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_data = Font(name="Segoe UI", size=10, color="000000")
        fill_even = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "HAIER - ESD MONITORING SYSTEM"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        ws['A2'] = "Report Type:"
        ws['A2'].font = font_meta_lbl
        ws['B2'] = range_type.upper()
        ws['B2'].font = font_meta_val
        
        ws['A3'] = "Generated At:"
        ws['A3'].font = font_meta_lbl
        ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['B3'].font = font_meta_val
        
        headers = ["S.No", "Station Code", "Transition Event", "Violation Start Time", "Violation End Time", "Duration (HH:MM:SS)", "Acknowledged"]
        for col_idx, text_val in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=text_val)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[5].height = 26
        
        row_idx = 6
        for idx, e in enumerate(events, start=1):
            c1 = ws.cell(row=row_idx, column=1, value=idx)
            c1.alignment = Alignment(horizontal="center")
            
            st_code = e.station.station_code.replace("ESD-STN-", "ST-") if e.station else ""
            c2 = ws.cell(row=row_idx, column=2, value=st_code)
            c2.alignment = Alignment(horizontal="center")
            
            c3 = ws.cell(row=row_idx, column=3, value=e.event_type)
            c3.alignment = Alignment(horizontal="center")
            
            # Start and End Time logic
            if e.event_type == "RESTORED" and e.duration_seconds is not None:
                start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                end_val = e.event_timestamp.replace(tzinfo=None)
            else:
                start_dt = e.event_timestamp
                end_val = "--"
                
            c4 = ws.cell(row=row_idx, column=4, value=start_dt.replace(tzinfo=None))
            c4.number_format = 'yyyy-mm-dd hh:mm:ss'
            c4.alignment = Alignment(horizontal="center")
            
            c5 = ws.cell(row=row_idx, column=5, value=end_val)
            if isinstance(end_val, datetime.datetime):
                c5.number_format = 'yyyy-mm-dd hh:mm:ss'
            c5.alignment = Alignment(horizontal="center")
            
            dur_str = "--"
            if e.duration_seconds is not None:
                hrs = e.duration_seconds // 3600
                mins = (e.duration_seconds % 3600) // 60
                secs = e.duration_seconds % 60
                dur_str = f"{hrs:02d}:{mins:02d}:{secs:02d}"
            c6 = ws.cell(row=row_idx, column=6, value=dur_str)
            c6.alignment = Alignment(horizontal="center")
            
            ack_str = "YES" if e.acknowledged else "NO"
            c7 = ws.cell(row=row_idx, column=7, value=ack_str)
            c7.alignment = Alignment(horizontal="center")
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_even
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime.datetime):
                        val_len = 19
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        ws.freeze_panes = 'A6'
        
        if row_idx > 6:
            ws.auto_filter.ref = f"A5:G{row_idx - 1}"
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename=esd_report_{range_type}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return response
    finally:
        db.remove_session()


@app.route("/search")
@login_required
def search():
    station_code = request.args.get("station_code")
    event_type = request.args.get("event_type")
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")
    
    performed = False
    results = []
    
    session_ref = db.get_session()
    try:
        if station_code or event_type or start_date_str or end_date_str:
            performed = True
            query = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(db.Station.id <= 20)
            
            if station_code:
                parsed_st = station_code.upper().strip()
                if parsed_st.startswith("ST-"):
                    parsed_st = parsed_st.replace("ST-", "ESD-STN-")
                query = query.filter(db.Station.station_code.ilike(f"%{parsed_st}%"))
                
            if event_type:
                query = query.filter(db.ESDEvent.event_type == event_type)
                
            if start_date_str:
                dt = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
                query = query.filter(db.ESDEvent.event_timestamp >= dt)
                
            if end_date_str:
                dt = datetime.datetime.strptime(end_date_str, "%Y-%m-%d") + datetime.timedelta(days=1)
                query = query.filter(db.ESDEvent.event_timestamp < dt)
                
            events = query.order_by(db.ESDEvent.event_timestamp.desc()).all()
            
            for e in events:
                if e.event_type == "RESTORED" and e.duration_seconds is not None:
                    start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                    end_dt = e.event_timestamp
                    start_str = start_dt.strftime('%Y-%m-%d %H:%M:%S')
                    end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    start_str = e.event_timestamp.strftime('%Y-%m-%d %H:%M:%S')
                    end_str = "Active" if (e.event_type == "VIOLATION" and e.station and e.station.current_state) else "--"
                
                results.append({
                    "station_id": e.station.id if e.station else 0,
                    "station_code": e.station.station_code.replace("ESD-STN-", "ST-") if e.station else "",
                    "description": e.station.description if e.station else "",
                    "event_type": e.event_type,
                    "start_time": start_str,
                    "end_time": end_str,
                    "duration_seconds": e.duration_seconds,
                    "acknowledged": e.acknowledged
                })
            
        return render_template(
            "search.html",
            active_page="search",
            station_code=station_code,
            event_type=event_type,
            start_date=start_date_str,
            end_date=end_date_str,
            performed=performed,
            results=results
        )
    except Exception as e:
        logging.error("Search query failed: %s", e)
        return render_template(
            "search.html",
            active_page="search",
            station_code=station_code,
            event_type=event_type,
            start_date=start_date_str,
            end_date=end_date_str,
            performed=performed,
            results=[],
            error=str(e)
        )
    finally:
        db.remove_session()


@app.route("/alarms")
@login_required
def alarms():
    session_ref = db.get_session()
    try:
        active_stations = session_ref.query(db.Station).filter_by(current_state=True).all()
        active_ids = [s.id for s in active_stations]
        
        active_events = []
        if active_ids:
            for sid in active_ids:
                ev = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).filter_by(station_id=sid, event_type="VIOLATION").order_by(db.ESDEvent.event_timestamp.desc()).first()
                if ev:
                    active_events.append(ev)
        
        active_events.sort(key=lambda x: x.event_timestamp, reverse=True)
        return render_template("alarms.html", active_page="alarms", active_events=active_events)
    finally:
        db.remove_session()


@app.route("/alarms/ack/<int:event_id>")
@login_required
def acknowledge_alarm(event_id):
    session_ref = db.get_session()
    try:
        ev = session_ref.get(db.ESDEvent, event_id)
        if ev and not ev.acknowledged:
            ev.acknowledged = True
            ev.acknowledged_by = session.get("user_id")
            ev.acknowledged_at = datetime.datetime.now(datetime.UTC).replace(tzinfo=None)
            session_ref.commit()
            
            today_start = datetime.datetime.now(datetime.UTC).replace(hour=0, minute=0, second=0, microsecond=0).replace(tzinfo=None)
            today_alerts_count = session_ref.query(db.ESDEvent).join(db.Station).filter(
                db.Station.id <= 20,
                db.ESDEvent.event_type == "VIOLATION",
                db.ESDEvent.event_timestamp >= today_start
            ).count()
            
            broadcaster.publish("esd_event", {
                "station_id": ev.station_id,
                "station_code": ev.station.station_code,
                "event_type": "ACKNOWLEDGED",
                "event_timestamp": ev.event_timestamp.isoformat(),
                "duration_seconds": ev.duration_seconds,
                "today_alerts_count": today_alerts_count,
                "violation_start": ev.event_timestamp.isoformat()
            })
    except Exception as e:
        session_ref.rollback()
        logging.error("Failed to acknowledge alarm: %s", e)
    finally:
        db.remove_session()
    return redirect(url_for("alarms"))


@app.route("/alarms/reset/<int:station_id>")
@login_required
def reset_alarm(station_id):
    session_ref = db.get_session()
    try:
        station = session_ref.get(db.Station, station_id)
        if station and station.current_state:
            event_service.record_transition(station, False, session_ref)
    except Exception as e:
        logging.error("Failed to reset alarm: %s", e)
    finally:
        db.remove_session()
    return redirect(url_for("alarms"))


@app.route("/settings")
@login_required
def settings_page():
    session_ref = db.get_session()
    try:
        stations = session_ref.query(db.Station).filter(db.Station.id <= 20).all()
        configs = session_ref.query(db.SystemConfig).all()
        config_map = {c.key: c.value for c in configs}
        success = request.args.get("success")
        return render_template("settings.html", active_page="settings", stations=stations, config_map=config_map, success=success)
    finally:
        db.remove_session()


@app.route("/settings/save", methods=["POST"])
@login_required
def settings_save():
    kc868_host = request.form.get("kc868_host")
    kc868_port = request.form.get("kc868_port", type=int)
    poll_interval_ms = request.form.get("poll_interval_ms", type=int)
    
    if not kc868_host or not kc868_port or not poll_interval_ms:
        return redirect(url_for("settings_page", error="All controller settings are required"))
        
    session_ref = db.get_session()
    try:
        h = session_ref.query(db.SystemConfig).filter_by(key="KC868_HOST").first()
        if h: h.value = kc868_host
        p = session_ref.query(db.SystemConfig).filter_by(key="KC868_PORT").first()
        if p: p.value = str(kc868_port)
        i = session_ref.query(db.SystemConfig).filter_by(key="POLL_INTERVAL_MS").first()
        if i: i.value = str(poll_interval_ms)
        
        stations = session_ref.query(db.Station).filter(db.Station.id <= 20).all()
        for station in stations:
            desc_val = request.form.get(f"station_desc_{station.id}")
            status_val = request.form.get(f"station_status_{station.id}")
            if desc_val is not None:
                station.description = desc_val
            if status_val is not None:
                station.status = status_val
                
        session_ref.commit()
        
        # Apply runtime settings dynamically
        cfg.KC868_HOST = kc868_host
        cfg.KC868_PORT = kc868_port
        cfg.POLL_INTERVAL_MS = poll_interval_ms
        
        driver.host = kc868_host
        driver.port = kc868_port
        poller.debounce = DebounceService(cfg.DEBOUNCE_CONSECUTIVE_POLLS)
        
        try:
            scheduler.reschedule_job('poll_job', trigger='interval', seconds=(poll_interval_ms / 1000))
        except Exception:
            pass
            
        return redirect(url_for("settings_page", success="true"))
    except Exception as e:
        session_ref.rollback()
        return redirect(url_for("settings_page", error=str(e)))
    finally:
        db.remove_session()


@app.route("/users")
@login_required
@role_required("ADMIN")
def users_page():
    session_ref = db.get_session()
    try:
        users = session_ref.query(db.User).all()
        success = request.args.get("success")
        error = request.args.get("error")
        return render_template("users.html", active_page="users", users=users, success=success, error=error)
    finally:
        db.remove_session()


@app.route("/users/add", methods=["POST"])
@login_required
@role_required("ADMIN")
def add_user():
    username = request.form.get("username")
    password = request.form.get("password")
    role = request.form.get("role")
    full_name = request.form.get("full_name")
    
    if not username or not password or not role:
        return redirect(url_for("users_page", error="Username, password, and role are required"))
        
    session_ref = db.get_session()
    try:
        existing = session_ref.query(db.User).filter_by(username=username).first()
        if existing:
            return redirect(url_for("users_page", error="Username already exists"))
            
        new_user = db.User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
            full_name=full_name,
            is_active=True
        )
        session_ref.add(new_user)
        session_ref.commit()
        return redirect(url_for("users_page", success=f"User {username} created successfully"))
    except Exception as e:
        session_ref.rollback()
        return redirect(url_for("users_page", error=str(e)))
    finally:
        db.remove_session()


@app.route("/users/delete/<int:user_id>")
@login_required
@role_required("ADMIN")
def delete_user(user_id):
    if user_id == session.get("user_id"):
        return redirect(url_for("users_page", error="You cannot delete your own account"))
        
    session_ref = db.get_session()
    try:
        user = session_ref.get(db.User, user_id)
        if not user:
            return redirect(url_for("users_page", error="User not found"))
            
        session_ref.delete(user)
        session_ref.commit()
        return redirect(url_for("users_page", success="User deleted successfully"))
    except Exception as e:
        session_ref.rollback()
        return redirect(url_for("users_page", error=str(e)))
    finally:
        db.remove_session()


@app.route("/users/change-password", methods=["POST"])
@login_required
@role_required("ADMIN")
def change_password():
    target_user_id = request.form.get("target_user_id", type=int)
    new_password = request.form.get("new_password")
    
    if not target_user_id or not new_password:
        return redirect(url_for("users_page", error="Missing user selection or password"))
        
    session_ref = db.get_session()
    try:
        user = session_ref.get(db.User, target_user_id)
        if not user:
            return redirect(url_for("users_page", error="User not found"))
            
        user.password_hash = generate_password_hash(new_password)
        session_ref.commit()
        return redirect(url_for("users_page", success=f"Password updated for {user.username}"))
    except Exception as e:
        session_ref.rollback()
        return redirect(url_for("users_page", error=str(e)))
    finally:
        db.remove_session()


# ==================================================================== #
# HTTP API Endpoint Services
# ==================================================================== #

@app.route("/api/v1/config", methods=["GET"])
@login_required
def api_get_config():
    return jsonify({
        "status": "success",
        "data": {
            "KC868_HOST": cfg.KC868_HOST,
            "KC868_PORT": cfg.KC868_PORT,
            "POLL_INTERVAL_MS": cfg.POLL_INTERVAL_MS,
            "TOTAL_KC868_CHANNELS": 20
        }
    })


def get_station_category(sid):
    if 1 <= sid <= 5:
        return "Backlight Line"
    elif 6 <= sid <= 10:
        return "Panel Line"
    elif 11 <= sid <= 20:
        return "General Assembly"
    return "Assembly Line"


@app.route("/api/v1/stations", methods=["GET"])
@login_required
def get_stations():
    session_ref = db.get_session()
    try:
        stations = session_ref.query(db.Station).filter(db.Station.id <= 20).all()
        data = []
        for s in stations:
            violation_start = None
            if s.current_state:
                latest_viol = session_ref.query(db.ESDEvent).filter_by(station_id=s.id, event_type="VIOLATION").order_by(db.ESDEvent.event_timestamp.desc()).first()
                if latest_viol:
                    violation_start = latest_viol.event_timestamp.isoformat()
            
            data.append({
                "id": s.id,
                "station_code": s.station_code,
                "kc868_channel": s.kc868_channel,
                "description": s.description,
                "category": get_station_category(s.id),
                "status": s.status,
                "current_state": s.current_state,
                "violation_start": violation_start
            })

        today_start = datetime.datetime.now(datetime.UTC).replace(hour=0, minute=0, second=0, microsecond=0).replace(tzinfo=None)
        today_alerts_count = session_ref.query(db.ESDEvent).join(db.Station).filter(
            db.Station.id <= 20,
            db.ESDEvent.event_type == "VIOLATION",
            db.ESDEvent.event_timestamp >= today_start
        ).count()

        return jsonify({
            "status": "success", 
            "data": data,
            "today_alerts_count": today_alerts_count
        })
    finally:
        db.remove_session()


@app.route("/api/v1/logs", methods=["GET"])
@login_required
def get_logs():
    session_ref = db.get_session()
    try:
        station_code = request.args.get("station_code")
        start_str = request.args.get("start")
        end_str = request.args.get("end")
        limit = request.args.get("limit", type=int)

        query = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(db.Station.id <= 20)

        if station_code:
            if "ST-" in station_code.upper():
                parsed = station_code.upper().replace("ST-", "ESD-STN-")
                query = query.filter(db.Station.station_code == parsed)
            else:
                query = query.filter(db.Station.station_code.ilike(f"%{station_code}%"))
        if start_str:
            dt = datetime.datetime.strptime(start_str, "%Y-%m-%d")
            query = query.filter(db.ESDEvent.event_timestamp >= dt)
        if end_str:
            dt = datetime.datetime.strptime(end_str, "%Y-%m-%d") + datetime.timedelta(days=1)
            query = query.filter(db.ESDEvent.event_timestamp < dt)

        query = query.order_by(db.ESDEvent.event_timestamp.desc())

        if limit:
            query = query.limit(limit)

        events = query.all()
        data = [{
            "id": e.id,
            "station_code": e.station.station_code,
            "station_description": e.station.description if e.station else "",
            "station_category": get_station_category(e.station.id) if e.station else "Assembly Line",
            "event_type": e.event_type,
            "event_timestamp": e.event_timestamp.isoformat(),
            "duration_seconds": e.duration_seconds,
            "acknowledged": e.acknowledged
        } for e in events]

        return jsonify({"status": "success", "data": data})
    finally:
        db.remove_session()


@app.route("/api/v1/export/csv", methods=["GET"])
@login_required
def export_csv():
    session_ref = db.get_session()
    try:
        events = session_ref.query(db.ESDEvent).options(joinedload(db.ESDEvent.station)).join(db.Station).filter(db.Station.id <= 20).order_by(db.ESDEvent.event_timestamp.desc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESD Audit Trail"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        fill_title = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
        font_meta_lbl = Font(name="Segoe UI", size=10, bold=True, color="748297")
        font_meta_val = Font(name="Segoe UI", size=10, bold=True, color="141B2D")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_data = Font(name="Segoe UI", size=10, color="000000")
        fill_even = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "HAIER - ESD MONITORING SYSTEM"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        ws['A2'] = "Report Type:"
        ws['A2'].font = font_meta_lbl
        ws['B2'] = "AUDIT TRAIL"
        ws['B2'].font = font_meta_val
        
        ws['A3'] = "Generated At:"
        ws['A3'].font = font_meta_lbl
        ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['B3'].font = font_meta_val
        
        headers = ["S.No", "Station Code", "Transition Event", "Violation Start Time", "Violation End Time", "Duration (HH:MM:SS)", "Acknowledged"]
        for col_idx, text_val in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=text_val)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[5].height = 26
        
        row_idx = 6
        for idx, e in enumerate(events, start=1):
            c1 = ws.cell(row=row_idx, column=1, value=idx)
            c1.alignment = Alignment(horizontal="center")
            
            st_code = e.station.station_code.replace("ESD-STN-", "ST-") if e.station else ""
            c2 = ws.cell(row=row_idx, column=2, value=st_code)
            c2.alignment = Alignment(horizontal="center")
            
            c3 = ws.cell(row=row_idx, column=3, value=e.event_type)
            c3.alignment = Alignment(horizontal="center")
            
            # Start and End Time logic
            if e.event_type == "RESTORED" and e.duration_seconds is not None:
                start_dt = e.event_timestamp - datetime.timedelta(seconds=e.duration_seconds)
                end_val = e.event_timestamp.replace(tzinfo=None)
            else:
                start_dt = e.event_timestamp
                end_val = "--"
                
            c4 = ws.cell(row=row_idx, column=4, value=start_dt.replace(tzinfo=None))
            c4.number_format = 'yyyy-mm-dd hh:mm:ss'
            c4.alignment = Alignment(horizontal="center")
            
            c5 = ws.cell(row=row_idx, column=5, value=end_val)
            if isinstance(end_val, datetime.datetime):
                c5.number_format = 'yyyy-mm-dd hh:mm:ss'
            c5.alignment = Alignment(horizontal="center")
            
            dur_str = "--"
            if e.duration_seconds is not None:
                hrs = e.duration_seconds // 3600
                mins = (e.duration_seconds % 3600) // 60
                secs = e.duration_seconds % 60
                dur_str = f"{hrs:02d}:{mins:02d}:{secs:02d}"
            c6 = ws.cell(row=row_idx, column=6, value=dur_str)
            c6.alignment = Alignment(horizontal="center")
            
            ack_str = "YES" if e.acknowledged else "NO"
            c7 = ws.cell(row=row_idx, column=7, value=ack_str)
            c7.alignment = Alignment(horizontal="center")
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_even
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime.datetime):
                        val_len = 19
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        ws.freeze_panes = 'A6'
        
        if row_idx > 6:
            ws.auto_filter.ref = f"A5:G{row_idx - 1}"
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename=esd_audit_trail_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return response
    finally:
        db.remove_session()


# ==================================================================== #
# Server-Sent Events (SSE) Stream Controller
# ==================================================================== #
@app.route("/stream")
@login_required
def stream():
    def event_generator():
        q = broadcaster.subscribe()
        try:
            session_ref = db.get_session()
            try:
                today_start = datetime.datetime.now(datetime.UTC).replace(hour=0, minute=0, second=0, microsecond=0).replace(tzinfo=None)
                today_alerts_count = session_ref.query(db.ESDEvent).join(db.Station).filter(
                    db.Station.id <= 20,
                    db.ESDEvent.event_type == "VIOLATION",
                    db.ESDEvent.event_timestamp >= today_start
                ).count()
                
                init_msg = json.dumps({"status": "ONLINE", "today_alerts_count": today_alerts_count})
                yield f"event: comm_status\ndata: {init_msg}\n\n"
            finally:
                db.remove_session()

            while True:
                message = q.get()
                yield message
        except GeneratorExit:
            pass
        finally:
            broadcaster.unsubscribe(q)

    return Response(stream_with_context(event_generator()), mimetype="text/event-stream")


if __name__ == "__main__":
    if cfg.ENV == "production":
        from waitress import serve
        print("Serving production build using Waitress on http://0.0.0.0:5000")
        serve(app, host="0.0.0.0", port=5000, threads=100)
    else:
        app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)